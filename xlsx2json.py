import openpyxl
import json 
import sys
#import pprint

def main():
    file_path = sys.argv[1]     # first arg is excel file 
    workbook = openpyxl.load_workbook(file_path)

    sheet_name = 'Sheet1'
    sheet = workbook[sheet_name]

    # make category
    category = dict()
    G_column = 8                # QnA Topic letter (in old version vas 7)
    H_column = G_column + 1     # QnA Topic name
    for category_id, category_name in sheet.iter_rows(min_row=2, min_col=G_column, max_col=H_column, values_only=True):
        #print(f'{row=}') 
        #print(f'{category_id=} {category_name=}')

        if not all((category_id, category_name)):
            break
            #raise ValueError(f'{category_id=} {category_name=}')
        else:
            category[category_id] = category_name
    #print(f'{category=}\n')


    # make all
    all_data = list()
    # max_row=3 only 2 first, for testing
    for qs, ql, links, _, _, _, category_id in sheet.iter_rows(min_row=2, max_col=7, values_only=True):#, max_row=3):
        #print(f'{qs=}, {ql=}, {links=}, {category_id=}\n')
        #break  # just for test
        
        if category_id == None:
            continue # this should be error

        #if not all((qs, ql, links, category_id)):
        if not all((qs, links, category_id)):
            break

        all_data.append((qs, ql, links.split(',')[0], category_id))

    # show when done
    all_json = json.dumps(all_data)
    #pprint.pprint(all_json)

    # save to HDD
    with open('categories.json', 'w') as file:
        json.dump(category, file, indent=2)

    with open('all_data.json', 'w') as file:
        json.dump(all_data, file, indent=2)



if __name__ == '__main__':
    main()
